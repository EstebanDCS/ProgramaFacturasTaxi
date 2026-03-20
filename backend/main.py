import os
import io
import re
import json
import zipfile
import shutil
import subprocess
from fastapi import FastAPI, Header, HTTPException, Depends, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from starlette.background import BackgroundTask
from pydantic import BaseModel
from typing import List, Optional
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
from supabase import create_client, Client

# ─── CONFIGURACIÓN ───────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("⚠️ ADVERTENCIA: Faltan SUPABASE_URL o SUPABASE_KEY")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY) if SUPABASE_URL and SUPABASE_KEY else None

app = FastAPI(title="Gestión Facturas API v2")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_methods=["*"], allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

BASE_DIR = os.path.dirname(__file__)
TEMP_DIR = os.path.join(BASE_DIR, "temp_files")

ADMIN_EMAILS = set(
    e.strip().lower()
    for e in os.environ.get("ADMIN_EMAILS", "estebandelka@gmail.com").split(",")
    if e.strip()
)


# ─── AUTENTICACIÓN ───────────────────────────────────────────────────────────
async def verificar_usuario(authorization: str = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="No hay sesión activa.")
    token = authorization.split(" ")[1]
    try:
        res = supabase.auth.get_user(token)
        if not res or not res.user:
            raise HTTPException(status_code=401, detail="Sesión inválida.")
        # Comprobar si el usuario está desactivado (admins son inmunes)
        email = (res.user.email or "").lower()
        if email not in ADMIN_EMAILS:
            try:
                settings = supabase.table("user_settings") \
                    .select("disabled").eq("user_id", res.user.id).execute()
                if settings.data and len(settings.data) > 0 and settings.data[0].get("disabled"):
                    raise HTTPException(status_code=403, detail="Cuenta desactivada por administrador.")
            except HTTPException:
                raise
            except:
                pass
        return res.user
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Token no válido: {e}")


async def verificar_admin(user=Depends(verificar_usuario)):
    if (user.email or "").lower() not in ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="Solo administradores.")
    return user


# ─── ACTIVITY LOG ────────────────────────────────────────────────────────────
def log_actividad(user_id: str, accion: str, meta: str = ""):
    """Registra una acción sin datos sensibles."""
    try:
        supabase.table("activity_log").insert({
            "user_id": user_id, "accion": accion, "meta": meta,
        }).execute()
    except Exception:
        pass  # No bloquear operaciones por fallo de log


# ─── MODELOS ─────────────────────────────────────────────────────────────────
class DatosTicket(BaseModel):
    numero_ticket: str
    importe: float
    contacto_metodo: str = ""
    fechas_solicitud: List[str] = []
    fechas_servicio: List[str] = []
    pasajeros: List[str] = []
    o_aeropuerto: bool = False
    o_buque: bool = False
    o_hotel: bool = False
    o_hotel_texto: str = ""
    o_otros: bool = False
    o_otros_texto: str = ""
    d_aeropuerto: bool = False
    d_buque: bool = False
    d_hotel: bool = False
    d_hotel_texto: str = ""
    d_hospital: bool = False
    d_hospital_texto: str = ""
    d_clinica: bool = False
    d_clinica_texto: str = ""
    d_inmigracion: bool = False
    d_otros: bool = False
    d_otros_texto: str = ""
    ida_vuelta: bool = False
    comentarios: str = ""


class DatosFactura(BaseModel):
    factura_numero: str
    barco: str
    tickets: List[DatosTicket]

    @property
    def importe_total(self) -> float:
        return sum(t.importe for t in self.tickets)


class DatosCliente(BaseModel):
    nombre: str = ""
    cif: str = ""
    direccion: str = ""
    email: str = ""
    telefono: str = ""


class DatosFacturaGenerica(BaseModel):
    numero_factura: str
    fecha: str = ""
    referencia: str = ""
    cliente: DatosCliente = DatosCliente()
    lineas: List[dict] = []
    notas: str = ""


# ─── UTILIDADES ──────────────────────────────────────────────────────────────
def compact_json(data: dict) -> str:
    def strip(obj):
        if isinstance(obj, dict):
            return {k: strip(v) for k, v in obj.items() if v not in (False, "", [], None)}
        if isinstance(obj, list):
            return [strip(i) for i in obj]
        return obj
    return json.dumps(strip(data), separators=(",", ":"))


def limpiar_temp(directorio: str):
    try:
        if os.path.exists(directorio):
            shutil.rmtree(directorio)
    except OSError:
        pass


def escribir_celda(ws, coord: str, valor):
    try:
        if type(ws[coord]).__name__ != "MergedCell":
            ws[coord] = valor
    except Exception:
        pass


def formatear_fechas(fechas: list, fmt_in="%Y-%m-%d", fmt_out="%d/%m/%Y") -> list:
    resultado = []
    for f in fechas:
        try:
            resultado.append(datetime.strptime(f, fmt_in).strftime(fmt_out))
        except ValueError:
            continue
    return resultado


def datos_registro(user_id: str, datos: DatosFactura, plantilla: str = "original") -> dict:
    """Construye dict para Supabase."""
    return {
        "user_id": user_id,
        "numero_factura": datos.factura_numero,
        "barco": datos.barco.upper(),
        "importe_total": datos.importe_total,
        "datos_json": compact_json(datos.dict()),
        "plantilla_nombre": plantilla,
    }


def datos_registro_cifrado(user_id: str, datos_json_cifrado: str, meta: dict) -> dict:
    """Para datos cifrados: el frontend envía el JSON ya cifrado."""
    return {
        "user_id": user_id,
        "numero_factura": meta.get("numero_factura", ""),
        "barco": meta.get("barco", ""),
        "importe_total": meta.get("importe_total", 0),
        "datos_json": datos_json_cifrado,
        "plantilla_nombre": meta.get("plantilla_nombre", "original"),
    }


# ─── PLANTILLAS: HELPERS ─────────────────────────────────────────────────────

def ticket_origen_texto(t: dict) -> str:
    partes = []
    if t.get("o_aeropuerto"): partes.append("Aeropuerto")
    if t.get("o_buque"): partes.append("Buque")
    if t.get("o_hotel"): partes.append(f"Hotel: {t.get('o_hotel_texto', '')}")
    if t.get("o_otros"): partes.append(t.get("o_otros_texto", "Otros"))
    return ", ".join(partes) or "—"


def ticket_destino_texto(t: dict) -> str:
    partes = []
    if t.get("d_aeropuerto"): partes.append("Aeropuerto")
    if t.get("d_buque"): partes.append("Buque")
    if t.get("d_hotel"): partes.append(f"Hotel: {t.get('d_hotel_texto', '')}")
    if t.get("d_hospital"): partes.append(f"Hospital: {t.get('d_hospital_texto', '')}")
    if t.get("d_clinica"): partes.append(f"Clínica: {t.get('d_clinica_texto', '')}")
    if t.get("d_inmigracion"): partes.append("Inmigración")
    if t.get("d_otros"): partes.append(t.get("d_otros_texto", "Otros"))
    return ", ".join(partes) or "—"


def escanear_tags_excel(wb) -> list:
    encontrados = set()
    patron = re.compile(r"\{\{[a-z_]+\}\}")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    encontrados.update(patron.findall(cell.value))
    return sorted(encontrados)


def rellenar_tags_hoja(ws, tags_map: dict):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                nuevo = cell.value
                for tag, valor in tags_map.items():
                    nuevo = nuevo.replace(tag, str(valor))
                if nuevo != cell.value:
                    try:
                        cell.value = float(nuevo)
                    except (ValueError, TypeError):
                        cell.value = nuevo


# ─── FÓRMULAS ────────────────────────────────────────────────────────────────
def evaluar_formula(formula: str, linea: dict) -> str:
    """Evalúa fórmula =campo1*campo2. Soporta +,-,*,/ y constantes."""
    if not formula or not formula.startswith("="):
        return formula
    expr = formula[1:].strip()
    for campo, valor in sorted(linea.items(), key=lambda x: -len(x[0])):
        try:
            num = float(valor) if valor else 0
        except (ValueError, TypeError):
            num = 0
        expr = expr.replace(campo, str(num))
    if re.match(r'^[\d\s\.\+\-\*\/\(\)]+$', expr):
        try:
            return f"{eval(expr):.2f}"
        except Exception:
            return "0.00"
    return "0.00"


def aplicar_formulas(lineas: list, columnas: list) -> list:
    """Aplica fórmulas a todas las líneas."""
    resultado = []
    for linea in lineas:
        ln = dict(linea)
        for col in columnas:
            if col.get("tipo") == "formula" and col.get("formula"):
                ln[col["campo"]] = evaluar_formula(col["formula"], ln)
        resultado.append(ln)
    return resultado


def generar_html_visual(config: dict, datos_dict: dict) -> str:
    """Genera HTML con CSS moderno para WeasyPrint."""
    emp = config.get("empresa", {})
    est = config.get("estilo", {})
    pag = config.get("pagina", {})
    cols = config.get("columnas", [
        {"nombre": "Descripción", "campo": "descripcion", "alineacion": "left"},
        {"nombre": "Importe", "campo": "importe", "alineacion": "right"},
    ])
    col_importe = config.get("columna_importe", "")
    if not col_importe and cols:
        for c in reversed(cols):
            if c.get("tipo") in ("moneda", "numero", "formula"):
                col_importe = c["campo"]; break
        if not col_importe:
            col_importe = cols[-1].get("campo", "importe")

    impuestos = config.get("impuestos", [{"nombre": "IVA", "porcentaje": 21}])
    pie = config.get("pie", {})
    color1 = est.get("color_primario", "#1a2e4a")
    color2 = est.get("color_secundario", "#0d6dfd")
    fuente = est.get("fuente", "Helvetica, Arial, sans-serif")
    tam = est.get("tam_fuente", 10)
    moneda = config.get("moneda", "\u20ac")
    titulo = config.get("titulo", "FACTURA")
    mg = pag.get("margenes", {})
    mt, mr, mb, ml = mg.get("top", 20), mg.get("right", 15), mg.get("bottom", 20), mg.get("left", 15)
    interlineado = pag.get("interlineado", 1.4)

    fecha = datos_dict.get("fecha", "") or datetime.now().strftime("%d/%m/%Y")
    cliente = datos_dict.get("cliente", {})
    lineas = datos_dict.get("lineas", [])
    notas = datos_dict.get("notas", "")
    lineas = aplicar_formulas(lineas, cols)

    # Logo
    logo_html = ""
    if emp.get("logo_url"):
        logo_html = f'<img src="{emp["logo_url"]}" class="logo" />'

    # Empresa
    emp_nombre = emp.get("nombre", "")
    emp_sub = []
    if emp.get("cif"): emp_sub.append(emp["cif"])
    if emp.get("direccion"): emp_sub.append(emp["direccion"])
    contacts = []
    if emp.get("telefono"): contacts.append(emp["telefono"])
    if emp.get("email"): contacts.append(emp["email"])
    if contacts: emp_sub.append(" · ".join(contacts))

    # Cliente
    cliente_html = ""
    if config.get("cliente", {}).get("mostrar") and cliente:
        cl = []
        if cliente.get("nombre"): cl.append(f'<strong>{cliente["nombre"]}</strong>')
        if cliente.get("cif"): cl.append(f'CIF: {cliente["cif"]}')
        if cliente.get("direccion"): cl.append(cliente["direccion"])
        if cliente.get("email"): cl.append(cliente["email"])
        if cl:
            cliente_html = f'''<div class="cliente-box">
<div class="cliente-label">FACTURAR A:</div>
{"<br>".join(cl)}
</div>'''

    # Columnas visibles
    cols_vis = [c for c in cols if not c.get("oculta")]

    # Header
    th_html = "".join(f'<th style="text-align:{c.get("alineacion","left")}">{c["nombre"]}</th>' for c in cols_vis)

    # Filas
    filas_html = ""
    for ri, linea in enumerate(lineas):
        cls = "row-alt" if ri % 2 == 1 else ""
        celdas = ""
        for c in cols_vis:
            val = linea.get(c.get("campo", ""), "")
            align = c.get("alineacion", "left")
            if c.get("tipo") in ("moneda", "formula") or c.get("campo") == col_importe:
                try: display = f'{float(val):,.2f} {moneda}'
                except: display = str(val)
                celdas += f'<td class="num" style="text-align:{align}"><strong>{display}</strong></td>'
            elif c.get("tipo") == "numero":
                try: display = f'{float(val):g}'
                except: display = str(val)
                celdas += f'<td style="text-align:{align}">{display}</td>'
            else:
                celdas += f'<td style="text-align:{align}">{str(val)}</td>'
        filas_html += f'<tr class="{cls}">{celdas}</tr>'

    # Totales
    subtotal = 0
    for linea in lineas:
        try: subtotal += float(linea.get(col_importe, 0))
        except: pass

    totales_html = ""
    total_final = subtotal
    if config.get("mostrar_desglose", True) and impuestos:
        totales_html += f'<tr><td class="tot-label">Subtotal</td><td class="tot-val">{subtotal:,.2f} {moneda}</td></tr>'
        for imp in impuestos:
            pct = float(imp.get("porcentaje", 0))
            monto = subtotal * pct / 100
            total_final = subtotal + monto
            totales_html += f'<tr><td class="tot-label">{imp.get("nombre","Impuesto")} ({pct:g}%)</td><td class="tot-val">{monto:,.2f} {moneda}</td></tr>'

    totales_html += f'<tr class="total-row"><td class="tot-label">TOTAL</td><td class="tot-val">{total_final:,.2f} {moneda}</td></tr>'

    ref_html = f'<div style="margin-top:4px"><strong>Ref:</strong> {datos_dict.get("referencia","")}</div>' if datos_dict.get("referencia") else ""
    notas_html = f'<p class="notas">{notas}</p>' if notas else ""
    pago_html = f'<p class="pago">{pie["datos_pago"]}</p>' if pie.get("mostrar_datos_pago") and pie.get("datos_pago") else ""

    # Hojas de detalle
    det = config.get("hoja_detalle", {})
    pags_det = ""
    if det.get("activar") and lineas:
        titulo_det = det.get("titulo", "Detalle")
        campos_det = det.get("campos", [c["campo"] for c in cols])
        nombres = {c["campo"]: c["nombre"] for c in cols}
        for idx, linea in enumerate(lineas, 1):
            filas_d = "".join(
                f'<tr><td class="det-label">{nombres.get(campo, campo.replace("_"," ").title())}</td><td class="det-val">{linea.get(campo,"")}</td></tr>'
                for campo in campos_det
            )
            pags_det += f'''
<div class="page-break"></div>
<table class="header-table">
<tr><td class="emp-col">{logo_html}<div class="emp-name" style="font-size:{tam+4}px">{emp_nombre}</div></td>
<td class="fac-col"><div class="fac-title" style="font-size:{tam+6}px">{titulo_det} #{idx}</div>
<div>Factura: {datos_dict.get("numero_factura","")}</div><div>{fecha}</div></td></tr>
</table>
<table class="det-table">{filas_d}</table>'''

    return f'''<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
@page {{
    size: {pag.get("size","A4")} {pag.get("orientation","portrait")};
    margin: {mt}mm {mr}mm {mb}mm {ml}mm;
}}
body {{
    font-family: {fuente};
    font-size: {tam}px;
    color: #333;
    margin: 0;
    padding: 0;
    line-height: {interlineado};
}}
.logo {{ max-height: 50px; max-width: 160px; display: block; margin-bottom: 8px; }}
.header-table {{ width: 100%; margin-bottom: 20px; }}
.header-table td {{ vertical-align: top; padding: 0; }}
.emp-col {{ width: 55%; }}
.fac-col {{ text-align: right; }}
.emp-name {{ font-size: {tam+8}px; font-weight: 800; color: {color1}; margin-bottom: 4px; }}
.emp-sub {{ font-size: {tam-1}px; color: #777; }}
.fac-title {{ font-size: {tam+14}px; font-weight: 900; color: {color2}; margin-bottom: 8px; }}
.cliente-box {{ background: #f5f6fa; padding: 10px 14px; margin: 14px 0; }}
.cliente-label {{ font-size: {tam-2}px; color: #999; font-weight: 700; margin-bottom: 4px; }}
.items-table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
.items-table th {{
    background-color: {color1};
    color: white;
    font-weight: 700;
    padding: 8px 10px;
    font-size: {tam}px;
    border: none;
}}
.items-table td {{
    padding: 7px 10px;
    border-bottom: 1px solid #e0e4e8;
}}
.items-table .row-alt td {{ background-color: #f8f9fb; }}
.totals-table {{ width: auto; margin-left: auto; margin-top: 16px; }}
.tot-label {{ padding: 4px 14px; text-align: right; color: #888; }}
.tot-val {{ padding: 4px 14px; text-align: right; font-weight: 600; min-width: 100px; }}
.total-row td {{
    border-top: 2px solid {color1};
    padding: 8px 14px;
    font-weight: 800;
    font-size: {tam+3}px;
    color: {color1};
    background-color: #f5f6fa;
}}
.notas {{ margin: 12px 0 4px; font-size: {tam-1}px; color: #555; font-style: italic; }}
.pago {{ font-size: {tam-1}px; color: #666; margin: 4px 0; }}
.footer {{ margin-top: 30px; padding-top: 10px; border-top: 1px solid #ddd; font-size: {tam-1}px; color: #999; }}
.page-break {{ page-break-before: always; }}
.det-table {{ width: 100%; border-collapse: collapse; }}
.det-label {{ padding: 8px 14px; font-weight: 600; color: {color1}; width: 40%; background-color: #f5f6fa; border-bottom: 1px solid #eaedf0; }}
.det-val {{ padding: 8px 14px; border-bottom: 1px solid #eaedf0; }}
</style></head><body>

<table class="header-table">
<tr>
<td class="emp-col">
{logo_html}
<div class="emp-name">{emp_nombre}</div>
{"<br>".join(f'<span class="emp-sub">{p}</span>' for p in emp_sub)}
</td>
<td class="fac-col">
<div class="fac-title">{titulo}</div>
<div><strong>Nº:</strong> {datos_dict.get("numero_factura","")}</div>
<div><strong>Fecha:</strong> {fecha}</div>
{ref_html}
</td>
</tr>
</table>

{cliente_html}

<table class="items-table">
<tr>{th_html}</tr>
{filas_html}
</table>

<table class="totals-table">
{totales_html}
</table>

{notas_html}

<div class="footer">
{pie.get("texto","")}{pago_html}
</div>

{pags_det}

</body></html>'''

def html_a_pdf(html_content: str) -> str:
    """Convierte HTML a PDF usando WeasyPrint."""
    from weasyprint import HTML as WeasyprintHTML
    os.makedirs(TEMP_DIR, exist_ok=True)
    pdf_path = os.path.join(TEMP_DIR, f"vis_{datetime.now().strftime('%H%M%S%f')}.pdf")
    try:
        WeasyprintHTML(string=html_content).write_pdf(pdf_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {e}")
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=500, detail="No se generó el PDF.")
    return pdf_path


def procesar_con_plantilla_excel(excel_bytes: bytes, datos_dict: dict) -> str:
    os.makedirs(TEMP_DIR, exist_ok=True)
    tmp_in = os.path.join(TEMP_DIR, f"tpl_{datetime.now().strftime('%H%M%S%f')}.xlsx")
    with open(tmp_in, "wb") as f:
        f.write(excel_bytes)
    wb = openpyxl.load_workbook(tmp_in)
    # Generic tag replacement on all sheets
    tags = {}
    for k, v in datos_dict.items():
        if isinstance(v, str):
            tags[f"{{{{{k}}}}}"] = v
        elif isinstance(v, (int, float)):
            tags[f"{{{{{k}}}}}"] = str(v)
    for ws in wb.worksheets:
        rellenar_tags_hoja(ws, tags)
    out_path = os.path.join(TEMP_DIR, f"gen_{datetime.now().strftime('%H%M%S%f')}.xlsx")
    wb.save(out_path)
    return out_path


# ─── GENERACIÓN DE EXCEL (PLANTILLA ORIGINAL) ────────────────────────────────
def crear_excel(datos_dict, nombre_base):
    template_path = os.path.join(BASE_DIR, "plantilla.xlsm")
    if not os.path.exists(template_path):
        raise HTTPException(status_code=410, detail="La plantilla original (.xlsm) ya no está disponible. Use el sistema de plantillas personalizadas.")
    os.makedirs(TEMP_DIR, exist_ok=True)
    name_xlsm = f"{nombre_base}.xlsm"
    path_salida = os.path.join(TEMP_DIR, name_xlsm)
    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws_factura = wb.worksheets[0]
    ws_template_bono = wb.worksheets[1] if len(wb.worksheets) > 1 else None
    esc = escribir_celda
    tickets = datos_dict.get("tickets", [])
    barco = datos_dict.get("barco", "").upper()

    esc(ws_factura, "E15", datos_dict.get("factura_numero", ""))
    esc(ws_factura, "C17", barco)

    todas_fechas = []
    for t in tickets:
        todas_fechas.extend(t.get("fechas_servicio", []))
    if todas_fechas:
        try:
            fechas_obj = [datetime.strptime(f, "%Y-%m-%d") for f in todas_fechas if f]
            esc(ws_factura, "F17", max(fechas_obj).strftime("%d/%m/%Y"))
        except Exception:
            esc(ws_factura, "F17", datetime.now().strftime("%d/%m/%Y"))
    else:
        esc(ws_factura, "F17", datetime.now().strftime("%d/%m/%Y"))

    numeros = [str(t.get("numero_ticket", "")).strip() for t in tickets if str(t.get("numero_ticket", "")).strip()]
    total_importe = sum(float(t.get("importe", 0)) for t in tickets)
    esc(ws_factura, "B21", " ".join(numeros))
    esc(ws_factura, "F21", round(total_importe, 2))
    try:
        ws_factura["B21"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    except Exception:
        pass

    base_imponible = total_importe / 1.10
    iva = total_importe - base_imponible
    esc(ws_factura, "F38", round(base_imponible, 2))
    esc(ws_factura, "F39", round(iva, 2))
    esc(ws_factura, "F40", round(total_importe, 2))

    MAPA_RECOGIDA = [
        ("o_aeropuerto", "C26", None), ("o_buque", "C27", None),
        ("o_hotel", "C28", ("o_hotel_texto", "E28")),
        ("o_otros", "C29", ("o_otros_texto", "D30")),
    ]
    MAPA_DESTINO = [
        ("d_aeropuerto", "C33", None), ("d_buque", "C34", None),
        ("d_hotel", "C35", ("d_hotel_texto", "E35")),
        ("d_hospital", "C36", ("d_hospital_texto", "E36")),
        ("d_clinica", "C37", ("d_clinica_texto", "E37")),
        ("d_inmigracion", "C38", None),
        ("d_otros", "C39", ("d_otros_texto", "D40")),
    ]

    if ws_template_bono and tickets:
        for i, t in enumerate(tickets):
            ws = ws_template_bono if i == 0 else wb.copy_worksheet(ws_template_bono)
            ws.title = f"Bono_{t.get('numero_ticket', i + 1)}"
            esc(ws, "E2", t.get("numero_ticket", ""))
            esc(ws, "E9", t.get("contacto_metodo", "").upper())
            esc(ws, "C12", ", ".join(formatear_fechas(t.get("fechas_solicitud", []))))
            esc(ws, "C14", ", ".join(formatear_fechas(t.get("fechas_servicio", []))))
            esc(ws, "C16", barco)
            esc(ws, "B19", ", ".join(t.get("pasajeros", [])))
            for campo, celda, texto_info in MAPA_RECOGIDA + MAPA_DESTINO:
                if t.get(campo):
                    esc(ws, celda, "X")
                    if texto_info:
                        esc(ws, texto_info[1], t.get(texto_info[0], ""))
            esc(ws, "C43" if t.get("ida_vuelta") else "D43", "X")
            esc(ws, "B46", t.get("comentarios", ""))
            esc(ws, "E51", float(t.get("importe", 0)))

    for sheet in wb.worksheets:
        try:
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 1
            sheet.page_setup.paperSize = 9
            if hasattr(sheet, "sheet_properties") and hasattr(sheet.sheet_properties, "pageSetUpPr"):
                sheet.sheet_properties.pageSetUpPr.fitToPage = True
        except Exception:
            pass

    wb.save(path_salida)
    return path_salida, name_xlsm


def procesar_descarga(datos_dict, formato):
    nombre_base = f"{datos_dict.get('barco', 'Factura').replace(' ', '_').upper()}_{datetime.now().strftime('%d_%m_%Y')}"
    path_xlsm, name_xlsm = crear_excel(datos_dict, nombre_base)
    cleanup = BackgroundTask(limpiar_temp, TEMP_DIR)

    if formato == "excel":
        return FileResponse(path_xlsm, filename=name_xlsm,
                            headers={"Content-Disposition": f"attachment; filename={name_xlsm}"},
                            background=cleanup)

    pdf_name = f"{nombre_base}.pdf"
    pdf_path = os.path.join(TEMP_DIR, pdf_name)
    lo_profile = os.path.join(TEMP_DIR, "lo_profile")
    try:
        subprocess.run(["libreoffice", f"-env:UserInstallation=file://{lo_profile}",
                        "--headless", "--convert-to", "pdf", "--outdir", TEMP_DIR, path_xlsm],
                       check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    except Exception:
        raise HTTPException(status_code=500, detail="Error al generar PDF.")

    if formato == "pdf":
        if os.path.exists(pdf_path):
            return FileResponse(pdf_path, filename=pdf_name,
                                headers={"Content-Disposition": f"attachment; filename={pdf_name}"},
                                background=cleanup)
        raise HTTPException(status_code=500, detail="No se pudo crear el PDF.")

    zip_name = f"{nombre_base}.zip"
    zip_path = os.path.join(TEMP_DIR, zip_name)
    with zipfile.ZipFile(zip_path, "w") as zipf:
        zipf.write(path_xlsm, arcname=name_xlsm)
        if os.path.exists(pdf_path):
            zipf.write(pdf_path, arcname=pdf_name)
    return FileResponse(zip_path, filename=zip_name,
                        headers={"Content-Disposition": f"attachment; filename={zip_name}"},
                        background=cleanup)


# ═══════════════════════════════════════════════════════════════════════════════
# RUTAS: FACTURAS (flujo original)
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/historial")
async def historial(user=Depends(verificar_usuario)):
    try:
        res = supabase.table("facturas") \
            .select("id, numero_factura, barco, importe_total, fecha_creacion, plantilla_nombre") \
            .eq("user_id", user.id).order("fecha_creacion", desc=True).execute()
    except Exception:
        # Fallback si plantilla_nombre no existe aún
        res = supabase.table("facturas") \
            .select("id, numero_factura, barco, importe_total, fecha_creacion") \
            .eq("user_id", user.id).order("fecha_creacion", desc=True).execute()
    return res.data


@app.post("/solo-guardar")
async def solo_guardar(datos: DatosFactura, user=Depends(verificar_usuario)):
    supabase.table("facturas").insert(datos_registro(user.id, datos)).execute()
    log_actividad(user.id, "guardar_factura", "plantilla=original")
    return {"msg": "✅ Datos guardados en la nube"}


@app.post("/generar")
async def generar(datos: DatosFactura, formato: str = "excel", user=Depends(verificar_usuario)):
    supabase.table("facturas").insert(datos_registro(user.id, datos)).execute()
    log_actividad(user.id, "generar_factura", f"formato={formato},plantilla=original")
    return procesar_descarga(datos.dict(), formato)


@app.get("/factura/{f_id}")
async def get_factura(f_id: int, user=Depends(verificar_usuario)):
    res = supabase.table("facturas").select("datos_json") \
        .eq("id", f_id).eq("user_id", user.id).single().execute()
    if not res.data:
        raise HTTPException(status_code=404)
    raw = res.data["datos_json"]
    # Si está cifrado (empieza con ENC:), devolver tal cual para que el cliente descifre
    if raw.startswith("ENC:"):
        return {"datos_json_cifrado": raw}
    return json.loads(raw)


@app.put("/actualizar/{f_id}")
async def actualizar(f_id: int, datos: DatosFactura, user=Depends(verificar_usuario)):
    registro = datos_registro(user.id, datos)
    del registro["user_id"]
    supabase.table("facturas").update(registro).eq("id", f_id).eq("user_id", user.id).execute()
    log_actividad(user.id, "actualizar_factura")
    return {"msg": "✅ Factura actualizada"}


@app.put("/actualizar-generar/{f_id}")
async def actualizar_generar(f_id: int, datos: DatosFactura, formato: str = "excel", user=Depends(verificar_usuario)):
    registro = datos_registro(user.id, datos)
    del registro["user_id"]
    supabase.table("facturas").update(registro).eq("id", f_id).eq("user_id", user.id).execute()
    log_actividad(user.id, "actualizar_generar", f"formato={formato}")
    return procesar_descarga(datos.dict(), formato)


@app.get("/re-descargar/{f_id}")
async def redescargar_file(f_id: int, formato: str = "excel", user=Depends(verificar_usuario)):
    res = supabase.table("facturas").select("datos_json") \
        .eq("id", f_id).eq("user_id", user.id).single().execute()
    if not res.data:
        raise HTTPException(status_code=404)
    raw = res.data["datos_json"]
    if raw.startswith("ENC:"):
        raise HTTPException(status_code=400, detail="Datos cifrados. Descifra en el cliente primero.")
    log_actividad(user.id, "redescargar", f"formato={formato}")
    return procesar_descarga(json.loads(raw), formato)


@app.delete("/eliminar-factura/{f_id}")
async def eliminar_factura(f_id: int, user=Depends(verificar_usuario)):
    supabase.table("facturas").delete().eq("id", f_id).eq("user_id", user.id).execute()
    log_actividad(user.id, "eliminar_factura")
    return {"msg": "Factura eliminada"}


@app.delete("/limpiar-historial")
async def limpiar(user=Depends(verificar_usuario)):
    supabase.table("facturas").delete().eq("user_id", user.id).execute()
    log_actividad(user.id, "limpiar_historial")
    return {"msg": "ok"}


# ═══════════════════════════════════════════════════════════════════════════════
# RUTAS: USER DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/user/stats")
async def user_stats(user=Depends(verificar_usuario)):
    """Estadísticas del usuario actual (ve sus propios datos)."""
    try:
        facturas = supabase.table("facturas") \
            .select("id, fecha_creacion, plantilla_nombre, importe_total") \
            .eq("user_id", user.id).execute().data or []
    except Exception:
        facturas = supabase.table("facturas") \
            .select("id, fecha_creacion, importe_total") \
            .eq("user_id", user.id).execute().data or []

    try:
        plantillas = supabase.table("plantillas") \
            .select("id, nombre, tipo") \
            .eq("user_id", user.id).execute().data or []
    except Exception:
        plantillas = []

    ahora = datetime.utcnow()
    facturas_mes = [f for f in facturas if f.get("fecha_creacion", "").startswith(ahora.strftime("%Y-%m"))]

    # Stats por plantilla
    por_plantilla = {}
    for f in facturas:
        p = f.get("plantilla_nombre", "original")
        if p not in por_plantilla:
            por_plantilla[p] = {"count": 0, "importe": 0}
        por_plantilla[p]["count"] += 1
        por_plantilla[p]["importe"] += f.get("importe_total", 0)

    # Facturas por mes (últimos 6)
    meses = {}
    for f in facturas:
        fc = f.get("fecha_creacion", "")
        if len(fc) >= 7:
            mk = fc[:7]
            meses[mk] = meses.get(mk, 0) + 1
    meses_ord = sorted(meses.items())[-6:]

    return {
        "total_facturas": len(facturas),
        "total_importe": round(sum(f.get("importe_total", 0) for f in facturas), 2),
        "facturas_este_mes": len(facturas_mes),
        "total_plantillas": len(plantillas),
        "por_plantilla": [{"nombre": k, **v} for k, v in por_plantilla.items()],
        "facturas_por_mes": [{"mes": m, "count": c} for m, c in meses_ord],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# RUTAS: PLANTILLAS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/plantillas")
async def listar_plantillas(user=Depends(verificar_usuario)):
    res = supabase.table("plantillas") \
        .select("id, nombre, tipo, config_json, es_publica, created_at, user_id") \
        .or_(f"user_id.eq.{user.id},es_publica.eq.true") \
        .order("created_at", desc=True).execute()
    return res.data


@app.get("/plantillas/{p_id}")
async def get_plantilla(p_id: int, user=Depends(verificar_usuario)):
    res = supabase.table("plantillas").select("*") \
        .eq("id", p_id).single().execute()
    if not res.data:
        raise HTTPException(status_code=404)
    p = res.data
    if p["user_id"] != user.id and not p.get("es_publica"):
        raise HTTPException(status_code=403)
    return p


@app.post("/plantillas/crear-visual")
async def crear_plantilla_visual(nombre: str = Form(...), config_json: str = Form(...), user=Depends(verificar_usuario)):
    try:
        json.loads(config_json)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="JSON inválido")
    res = supabase.table("plantillas").insert({
        "user_id": user.id, "nombre": nombre, "tipo": "visual", "config_json": config_json,
    }).execute()
    log_actividad(user.id, "crear_plantilla", f"tipo=visual,nombre={nombre}")
    return {"msg": "✅ Plantilla creada", "id": res.data[0]["id"] if res.data else None}


@app.post("/plantillas/upload-excel")
async def upload_excel_template(file: UploadFile = File(...), nombre: str = Form(...), config_json: str = Form("{}"), user=Depends(verificar_usuario)):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Solo .xlsx o .xlsm")
    contenido = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido))
        tags = escanear_tags_excel(wb)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo Excel: {e}")
    ext = os.path.splitext(file.filename)[1]
    storage_path = f"{user.id}/{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
    try:
        supabase.storage.from_("plantillas").upload(path=storage_path, file=contenido,
            file_options={"content-type": file.content_type or "application/octet-stream"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error subiendo: {e}")
    cfg = json.loads(config_json) if config_json else {}
    cfg["tags_detectados"] = tags
    cfg["filename_original"] = file.filename
    res = supabase.table("plantillas").insert({
        "user_id": user.id, "nombre": nombre, "tipo": "excel",
        "config_json": json.dumps(cfg, separators=(",", ":")), "excel_path": storage_path,
    }).execute()
    log_actividad(user.id, "crear_plantilla", f"tipo=excel,nombre={nombre}")
    return {"msg": "✅ Plantilla subida", "id": res.data[0]["id"] if res.data else None, "tags_detectados": tags}


@app.put("/plantillas/{p_id}")
async def actualizar_plantilla(p_id: int, nombre: str = Form(...), config_json: str = Form(...), user=Depends(verificar_usuario)):
    supabase.table("plantillas").update({"nombre": nombre, "config_json": config_json}) \
        .eq("id", p_id).eq("user_id", user.id).execute()
    return {"msg": "✅ Actualizada"}


@app.delete("/plantillas/{p_id}")
async def eliminar_plantilla(p_id: int, user=Depends(verificar_usuario)):
    res = supabase.table("plantillas").select("excel_path") \
        .eq("id", p_id).eq("user_id", user.id).single().execute()
    if not res.data:
        raise HTTPException(status_code=404)
    if res.data.get("excel_path"):
        try:
            supabase.storage.from_("plantillas").remove([res.data["excel_path"]])
        except Exception:
            pass
    supabase.table("plantillas").delete().eq("id", p_id).eq("user_id", user.id).execute()
    log_actividad(user.id, "eliminar_plantilla")
    return {"msg": "Eliminada"}


@app.post("/plantillas/{p_id}/preview")
async def preview_plantilla(p_id: int, user=Depends(verificar_usuario)):
    res = supabase.table("plantillas").select("*").eq("id", p_id).single().execute()
    if not res.data:
        raise HTTPException(status_code=404)
    p = res.data
    if p["user_id"] != user.id and not p.get("es_publica"):
        raise HTTPException(status_code=403)
    ejemplo = {
        "numero_factura": "2026-0001", "fecha": "19/03/2026", "referencia": "REF-EJEMPLO",
        "cliente": {"nombre": "Empresa Ejemplo S.L.", "cif": "B12345678", "direccion": "Calle Mayor 10, Madrid"},
        "lineas": [
            {"descripcion": "Servicio profesional", "cantidad": "10", "precio": "50.00", "horas": "10", "tarifa": "50.00", "importe": "500.00", "total": "500.00"},
            {"descripcion": "Material fungible", "cantidad": "1", "precio": "30.00", "horas": "2", "tarifa": "30.00", "importe": "30.00", "total": "30.00"},
        ],
        "notas": "Pago a 30 días.",
    }
    cleanup = BackgroundTask(limpiar_temp, TEMP_DIR)
    if p["tipo"] == "visual":
        config = json.loads(p["config_json"])
        html = generar_html_visual(config, ejemplo)
        pdf_path = html_a_pdf(html)
        return FileResponse(pdf_path, filename="preview.pdf",
                            headers={"Content-Disposition": "inline; filename=preview.pdf"}, background=cleanup)
    elif p["tipo"] == "excel":
        if not p.get("excel_path"):
            raise HTTPException(status_code=400)
        excel_bytes = supabase.storage.from_("plantillas").download(p["excel_path"])
        out = procesar_con_plantilla_excel(excel_bytes, ejemplo)
        return FileResponse(out, filename="preview.xlsx", background=cleanup)


@app.get("/health")
async def health():
    """Health check real — verifica que Supabase responde."""
    try:
        supabase.table("plantillas").select("id").limit(1).execute()
        return {"status": "ok"}
    except Exception:
        raise HTTPException(status_code=503, detail="Servidor iniciando...")


@app.post("/guardar-con-plantilla")
async def guardar_con_plantilla(datos: DatosFacturaGenerica, plantilla_id: Optional[int] = None, user=Depends(verificar_usuario)):
    """Guarda factura en la nube sin generar PDF."""
    if not plantilla_id:
        raise HTTPException(status_code=400, detail="Se requiere plantilla_id")
    res = supabase.table("plantillas").select("nombre").eq("id", plantilla_id).single().execute()
    plantilla_nombre = res.data["nombre"] if res.data else "custom"

    # Calcular importe total desde las líneas
    config = {}
    try:
        pr = supabase.table("plantillas").select("config_json").eq("id", plantilla_id).single().execute()
        if pr.data:
            config = json.loads(pr.data.get("config_json", "{}"))
    except Exception:
        pass

    cols = config.get("columnas", [])
    col_importe = config.get("columna_importe", "")
    if not col_importe and cols:
        for c in reversed(cols):
            if c.get("tipo") in ("moneda", "numero", "formula"):
                col_importe = c["campo"]
                break
    importe_total = 0
    for linea in datos.lineas:
        try:
            importe_total += float(linea.get(col_importe, 0))
        except (ValueError, TypeError):
            pass

    registro = {
        "user_id": user.id,
        "numero_factura": datos.numero_factura,
        "barco": datos.referencia or plantilla_nombre,
        "importe_total": round(importe_total, 2),
        "datos_json": compact_json(datos.dict()),
        "plantilla_nombre": plantilla_nombre,
    }
    supabase.table("facturas").insert(registro).execute()
    log_actividad(user.id, "guardar_con_plantilla", f"plantilla={plantilla_nombre}")
    return {"msg": "Factura guardada en la nube"}


@app.post("/generar-con-plantilla")
async def generar_con_plantilla(datos: DatosFacturaGenerica, formato: str = "pdf", plantilla_id: Optional[int] = None, user=Depends(verificar_usuario)):
    if not plantilla_id:
        raise HTTPException(status_code=400, detail="Se requiere plantilla_id")
    res = supabase.table("plantillas").select("*").eq("id", plantilla_id).single().execute()
    if not res.data:
        raise HTTPException(status_code=404)
    p = res.data
    if p["user_id"] != user.id and not p.get("es_publica"):
        raise HTTPException(status_code=403)

    nombre_base = f"Factura_{datos.numero_factura.replace(' ', '_')}_{datetime.now().strftime('%d_%m_%Y')}"
    cleanup = BackgroundTask(limpiar_temp, TEMP_DIR)
    log_actividad(user.id, "generar_con_plantilla", f"plantilla={p['nombre']},formato={formato}")

    if p["tipo"] == "visual":
        config = json.loads(p["config_json"])
        html = generar_html_visual(config, datos.dict())
        if formato == "html":
            return Response(content=html, media_type="text/html",
                            headers={"Content-Disposition": f"attachment; filename={nombre_base}.html"})
        try:
            pdf_path = html_a_pdf(html)
            pdf_name = f"{nombre_base}.pdf"
            final_path = os.path.join(TEMP_DIR, pdf_name)
            shutil.move(pdf_path, final_path)
            return FileResponse(final_path, filename=pdf_name,
                                headers={"Content-Disposition": f"attachment; filename={pdf_name}"}, background=cleanup)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error generando PDF: {e}. Prueba con formato=html.")
    elif p["tipo"] == "excel":
        if not p.get("excel_path"):
            raise HTTPException(status_code=400)
        excel_bytes = supabase.storage.from_("plantillas").download(p["excel_path"])
        out = procesar_con_plantilla_excel(excel_bytes, datos.dict())
        xlsx_name = f"{nombre_base}.xlsx"
        final_path = os.path.join(TEMP_DIR, xlsx_name)
        shutil.move(out, final_path)
        if formato == "excel":
            return FileResponse(final_path, filename=xlsx_name,
                                headers={"Content-Disposition": f"attachment; filename={xlsx_name}"}, background=cleanup)
        try:
            lo_profile = os.path.join(TEMP_DIR, "lo_profile_tpl")
            subprocess.run(["libreoffice", f"-env:UserInstallation=file://{lo_profile}",
                            "--headless", "--convert-to", "pdf", "--outdir", TEMP_DIR, final_path],
                           check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=60)
            pdf_name = f"{nombre_base}.pdf"
            pdf_path = os.path.join(TEMP_DIR, pdf_name)
            return FileResponse(pdf_path, filename=pdf_name,
                                headers={"Content-Disposition": f"attachment; filename={pdf_name}"}, background=cleanup)
        except Exception as e:
            return FileResponse(final_path, filename=xlsx_name,
                                headers={"Content-Disposition": f"attachment; filename={xlsx_name}"}, background=cleanup)


# ═══════════════════════════════════════════════════════════════════════════════
# RUTAS: ADMIN (sin datos financieros — privacidad total)
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/admin/check")
async def admin_check(user=Depends(verificar_usuario)):
    return {"is_admin": (user.email or "").lower() in ADMIN_EMAILS}


@app.get("/admin/stats")
async def admin_stats(user=Depends(verificar_admin)):
    """Estadísticas de USO de la plataforma (sin datos financieros)."""
    facturas = supabase.table("facturas") \
        .select("id, user_id, fecha_creacion, plantilla_nombre") \
        .order("fecha_creacion", desc=True).execute().data or []

    plantillas = supabase.table("plantillas") \
        .select("id, user_id, tipo, created_at").execute().data or []

    ahora = datetime.utcnow()
    usuarios_unicos = len(set(f.get("user_id", "") for f in facturas))
    facturas_hoy = [f for f in facturas if f.get("fecha_creacion", "").startswith(ahora.strftime("%Y-%m-%d"))]
    facturas_mes = [f for f in facturas if f.get("fecha_creacion", "").startswith(ahora.strftime("%Y-%m"))]

    # Uso por plantilla
    uso_plantillas = {}
    for f in facturas:
        p = f.get("plantilla_nombre", "original")
        uso_plantillas[p] = uso_plantillas.get(p, 0) + 1

    # Facturas por día (últimos 14 días)
    dias = {}
    for f in facturas:
        fc = f.get("fecha_creacion", "")
        if len(fc) >= 10:
            dias[fc[:10]] = dias.get(fc[:10], 0) + 1
    dias_ord = sorted(dias.items())[-14:]

    return {
        "total_facturas": len(facturas),
        "total_plantillas": len(plantillas),
        "usuarios_unicos": usuarios_unicos,
        "facturas_hoy": len(facturas_hoy),
        "facturas_este_mes": len(facturas_mes),
        "uso_plantillas": [{"nombre": k, "count": v} for k, v in sorted(uso_plantillas.items(), key=lambda x: -x[1])],
        "facturas_por_dia": [{"dia": d, "count": c} for d, c in dias_ord],
    }


@app.get("/admin/users")
async def admin_users(user=Depends(verificar_admin)):
    """Lista de usuarios con estadísticas de uso (sin datos personales de sus facturas)."""
    facturas = supabase.table("facturas") \
        .select("user_id, fecha_creacion").execute().data or []

    plantillas = supabase.table("plantillas") \
        .select("user_id").execute().data or []

    settings = supabase.table("user_settings") \
        .select("user_id, disabled").execute().data or []
    disabled_map = {s["user_id"]: s["disabled"] for s in settings}

    # Agregar por usuario
    users = {}
    for f in facturas:
        uid = f.get("user_id", "")
        if uid not in users:
            users[uid] = {"user_id": uid, "facturas": 0, "plantillas": 0, "ultima_actividad": "", "disabled": disabled_map.get(uid, False)}
        users[uid]["facturas"] += 1
        fc = f.get("fecha_creacion", "")
        if fc > users[uid]["ultima_actividad"]:
            users[uid]["ultima_actividad"] = fc

    for p in plantillas:
        uid = p.get("user_id", "")
        if uid in users:
            users[uid]["plantillas"] += 1
        else:
            users[uid] = {"user_id": uid, "facturas": 0, "plantillas": 1, "ultima_actividad": "", "disabled": disabled_map.get(uid, False)}

    return sorted(users.values(), key=lambda x: x["ultima_actividad"], reverse=True)


@app.put("/admin/users/{uid}/toggle")
async def admin_toggle_user(uid: str, user=Depends(verificar_admin)):
    """Activar/desactivar un usuario. No se puede desactivar a un admin."""
    # Protección: un admin no puede desactivarse a sí mismo
    if uid == user.id:
        raise HTTPException(status_code=400, detail="No puedes desactivar tu propia cuenta de administrador.")
    existing = supabase.table("user_settings") \
        .select("disabled").eq("user_id", uid).execute().data
    if existing:
        new_state = not existing[0].get("disabled", False)
        supabase.table("user_settings").update({"disabled": new_state, "updated_at": datetime.utcnow().isoformat()}) \
            .eq("user_id", uid).execute()
    else:
        new_state = True
        supabase.table("user_settings").insert({"user_id": uid, "disabled": True}).execute()
    log_actividad(user.id, "admin_toggle_user", f"target={uid[:8]},disabled={new_state}")
    return {"msg": f"Usuario {'desactivado' if new_state else 'activado'}", "disabled": new_state}


@app.get("/admin/activity")
async def admin_activity(user=Depends(verificar_admin), limit: int = 50):
    """Últimas acciones de todos los usuarios (sin contenido sensible)."""
    res = supabase.table("activity_log") \
        .select("id, user_id, accion, meta, created_at") \
        .order("created_at", desc=True) \
        .limit(limit).execute()
    return res.data


@app.get("/admin/system")
async def admin_system(user=Depends(verificar_admin)):
    """Estado del sistema."""
    # Contar registros
    facturas_count = len(supabase.table("facturas").select("id").execute().data or [])
    plantillas_count = len(supabase.table("plantillas").select("id").execute().data or [])
    logs_count = len(supabase.table("activity_log").select("id").execute().data or [])

    # Espacio temp
    temp_size = 0
    if os.path.exists(TEMP_DIR):
        for f in os.listdir(TEMP_DIR):
            fp = os.path.join(TEMP_DIR, f)
            if os.path.isfile(fp):
                temp_size += os.path.getsize(fp)

    return {
        "facturas_en_bd": facturas_count,
        "plantillas_en_bd": plantillas_count,
        "logs_en_bd": logs_count,
        "temp_files_mb": round(temp_size / 1024 / 1024, 2),
        "admin_emails": list(ADMIN_EMAILS),
        "server_time": datetime.utcnow().isoformat(),
    }